from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.parsers.models import ParsedDocument


HEADER_HINTS = ("项目", "名称", "区", "地点", "建设", "投资", "面积", "日期", "年度")
OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def parse_excel(
    content_or_path: bytes | str | Path,
    *,
    url: str,
    source_id: str = "",
    content_hash: str = "",
    fetched_at: str = "",
    raw_path: str = "",
) -> ParsedDocument:
    suffix_hint = excel_suffix_hint(raw_path=raw_path, url=url)
    if suffix_hint == ".xls" or looks_like_legacy_xls(content_or_path):
        return parse_legacy_xls(
            content_or_path,
            url=url,
            source_id=source_id,
            content_hash=content_hash,
            fetched_at=fetched_at,
            raw_path=raw_path,
        )
    stream: BytesIO | str | Path
    if isinstance(content_or_path, bytes):
        stream = BytesIO(content_or_path)
    else:
        stream = content_or_path
    workbook = load_workbook(stream, data_only=True, read_only=True)
    try:
        sheet_values = [
            (sheet.title, [[_cell_to_str(cell) for cell in row] for row in sheet.iter_rows(values_only=True)])
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()
    return parsed_from_sheet_values(
        source_id=source_id,
        url=url,
        title=document_title(raw_path=raw_path, url=url, fallback="Excel document"),
        content_hash=content_hash,
        fetched_at=fetched_at,
        raw_path=raw_path,
        sheet_values=sheet_values,
        metadata_extra={},
    )


def parse_legacy_xls(
    content_or_path: bytes | str | Path,
    *,
    url: str,
    source_id: str = "",
    content_hash: str = "",
    fetched_at: str = "",
    raw_path: str = "",
) -> ParsedDocument:
    try:
        import pandas as pd

        stream: BytesIO | str | Path
        stream = BytesIO(content_or_path) if isinstance(content_or_path, bytes) else content_or_path
        sheets = pd.read_excel(stream, sheet_name=None, header=None, dtype=str, engine="xlrd")
    except Exception as exc:
        return ParsedDocument(
            source_id=source_id,
            url=url,
            title=document_title(raw_path=raw_path, url=url, fallback="Legacy Excel document"),
            text=f"Legacy .xls could not be parsed. Install xlrd>=2.0 or convert to .xlsx. Error: {exc}",
            content_hash=content_hash,
            fetched_at=fetched_at,
            raw_path=raw_path,
            parser="excel",
            metadata={"unsupported": ".xls", "row_count": 0, "error": str(exc)},
        )
    sheet_values = []
    for sheet_name, frame in sheets.items():
        frame = frame.fillna("")
        values = [[_cell_to_str(value) for value in row] for row in frame.to_numpy().tolist()]
        sheet_values.append((str(sheet_name), values))
    return parsed_from_sheet_values(
        source_id=source_id,
        url=url,
        title=document_title(raw_path=raw_path, url=url, fallback="Legacy Excel document"),
        content_hash=content_hash,
        fetched_at=fetched_at,
        raw_path=raw_path,
        sheet_values=sheet_values,
        metadata_extra={"legacy_xls": True},
    )


def parsed_from_sheet_values(
    *,
    source_id: str,
    url: str,
    title: str,
    content_hash: str,
    fetched_at: str,
    raw_path: str,
    sheet_values: list[tuple[str, list[list[str]]]],
    metadata_extra: dict[str, Any],
) -> ParsedDocument:
    all_rows: list[dict[str, Any]] = []
    text_blocks: list[str] = []
    tables: list[list[list[str]]] = []

    for sheet_name, values in sheet_values:
        values = [trim_trailing_empty(row) for row in values if any(cell for cell in row)]
        if not values:
            continue
        header_idx = detect_header(values)
        headers = dedupe_headers(values[header_idx])
        rows = values[header_idx + 1 :]
        tables.append([headers] + rows[:100])
        text_blocks.append(f"## sheet: {sheet_name}")
        text_blocks.append(markdown_table([headers] + rows[:30]))
        for row in rows:
            if not any(row):
                continue
            mapped = {headers[idx] if idx < len(headers) else f"col_{idx + 1}": row[idx] if idx < len(row) else "" for idx in range(max(len(headers), len(row)))}
            mapped = {key: value for key, value in mapped.items() if key and value != ""}
            if mapped:
                mapped["_sheet"] = sheet_name
                all_rows.append(mapped)
                text_blocks.append(" | ".join(str(value) for value in mapped.values() if value != ""))

    return ParsedDocument(
        source_id=source_id,
        url=url,
        title=title,
        text="\n".join(text_blocks),
        rows=all_rows,
        tables=tables,
        content_hash=content_hash,
        fetched_at=fetched_at,
        raw_path=raw_path,
        parser="excel",
        metadata={"sheet_count": len(sheet_values), "row_count": len(all_rows), **metadata_extra},
    )


def detect_header(rows: list[list[str]]) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:12]):
        non_empty = [cell for cell in row if cell]
        hint_score = sum(1 for cell in non_empty if any(hint in cell for hint in HEADER_HINTS))
        score = len(non_empty) + hint_score * 2
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def dedupe_headers(headers: list[str]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}
    for idx, header in enumerate(headers):
        value = header.strip() or f"col_{idx + 1}"
        count = seen.get(value, 0)
        seen[value] = count + 1
        result.append(value if count == 0 else f"{value}_{count + 1}")
    return result


def markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    padded = [row + [""] * (width - len(row)) for row in rows]
    header = padded[0]
    separator = ["---"] * width
    body = padded[1:]
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(separator) + " |"]
    lines.extend("| " + " | ".join(row) + " |" for row in body)
    return "\n".join(lines)


def trim_trailing_empty(row: list[str]) -> list[str]:
    while row and row[-1] == "":
        row.pop()
    return row


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()[:10]
        except Exception:
            pass
    return str(value).strip()


def excel_suffix_hint(*, raw_path: str, url: str) -> str:
    if raw_path:
        return Path(raw_path).suffix.lower()
    from urllib.parse import unquote, urlparse

    parsed = urlparse(url)
    candidate = parsed.path or parsed.netloc or url
    return Path(unquote(candidate)).suffix.lower()


def document_title(*, raw_path: str, url: str, fallback: str) -> str:
    if raw_path:
        return Path(raw_path).name or fallback
    from urllib.parse import unquote, urlparse

    parsed = urlparse(url)
    candidate = parsed.path or parsed.netloc or url
    return Path(unquote(candidate)).name or fallback


def looks_like_legacy_xls(content_or_path: bytes | str | Path) -> bool:
    if isinstance(content_or_path, bytes):
        return content_or_path.startswith(OLE2_SIGNATURE)
    try:
        with Path(content_or_path).open("rb") as handle:
            return handle.read(len(OLE2_SIGNATURE)) == OLE2_SIGNATURE
    except OSError:
        return False
